from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import Callable, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel as excel_date
from openpyxl.worksheet.worksheet import Worksheet

class PipelineError(Exception):
    """Custom exception for consolidation pipeline errors."""

    def __init__(self, message: str, phase: Optional[int] = None):
        super().__init__(message)
        self.phase = phase


def run_workbook_pipeline(
    dms_bytes: bytes,
    rep_bytes: bytes,
    main_bytes: bytes,
    progress_callback: Callable[..., None],
) -> Tuple[BytesIO, str]:
    """Execute all pipeline phases and return the consolidated workbook buffer."""

    dms_wb = load_workbook(filename=BytesIO(dms_bytes), data_only=False, keep_links=True)
    rep_wb = load_workbook(filename=BytesIO(rep_bytes), data_only=False, keep_links=True)
    main_wb = load_workbook(filename=BytesIO(main_bytes), data_only=False, keep_links=True)

    _phase_one_normalize_dms(dms_wb, main_wb, progress_callback)
    phase_two_context = _phase_two_merge_rep(rep_wb, main_wb, progress_callback)
    _phase_three_update_main(main_wb, phase_two_context, progress_callback)

    _report(progress_callback, 4, status="done", percent=100, message="Phase 4: Completed, ready for human review.")

    output = BytesIO()
    main_wb.save(output)
    output.seek(0)
    filename = f"Main_Consolidated_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename


def _phase_one_normalize_dms(dms_wb, main_wb, progress_callback):
    sheet = dms_wb['DMS Dump'] if 'DMS Dump' in dms_wb.sheetnames else dms_wb.worksheets[0]
    headers, header_map, header_row = _build_header_index(sheet, required_headers=['Device_ID'])
    device_idx = header_map.get('device_id')
    if not device_idx:
        _report(progress_callback, 1, status='error', message="Phase 1 error: column 'Device_ID' not found in DMS file")
        raise PipelineError("Phase 1 error: column 'Device_ID' not found in DMS file", phase=1)

    data_rows = list(sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True))
    total = len(data_rows)
    _report(progress_callback, 1, status='running', total_rows=total, processed_rows=0,
            message='Phase 1: Normalizing Device_ID values…')

    dms_sheet_main = main_wb['DMS Dump'] if 'DMS Dump' in main_wb.sheetnames else main_wb.create_sheet(title='DMS Dump')

    # Ensure header row aligns with incoming file
    if dms_sheet_main.max_row == 0:
        dms_sheet_main.append(headers)
    else:
        for col_idx, header in enumerate(headers, start=1):
            dms_sheet_main.cell(row=1, column=col_idx, value=header)
        if dms_sheet_main.max_column > len(headers):
            for extra_col in range(len(headers) + 1, dms_sheet_main.max_column + 1):
                dms_sheet_main.cell(row=1, column=extra_col, value=None)

    if dms_sheet_main.max_row > 1:
        dms_sheet_main.delete_rows(2, dms_sheet_main.max_row - 1)

    for idx, row in enumerate(data_rows, start=1):
        normalized_row = list(row)
        device_value = normalized_row[device_idx - 1]
        normalized_row[device_idx - 1] = _normalize_device_id(device_value)
        dms_sheet_main.append(normalized_row)
        if idx % 50 == 0 or idx == total:
            _report(progress_callback, 1, processed_rows=idx, total_rows=total,
                    message=f"Phase 1: DMS normalization – {idx:,} / {total:,} rows")

    _report(progress_callback, 1, status='done', processed_rows=total, total_rows=total,
            message='Phase 1 complete – DMS Dump sheet refreshed inside MAIN.')


def _phase_two_merge_rep(rep_wb, main_wb, progress_callback):
    sheet = rep_wb.worksheets[0]
    headers, header_map, header_row = _build_header_index(sheet, required_headers=['Begin Journey Date'])
    begin_key = header_map.get('begin journey date')
    if not begin_key:
        _report(progress_callback, 2, status='error', message="Phase 2 error: column 'Begin Journey Date' not found in repJourney file")
        raise PipelineError("Phase 2 error: column 'Begin Journey Date' not found in repJourney file", phase=2)

    rep_rows: List[Dict[str, object]] = []
    header_keys = [ _normalize_header(h) for h in headers ]
    for values in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True):
        if not any(values):
            continue
        row_dict = { header_keys[i]: values[i] for i in range(len(header_keys)) if header_keys[i] }
        rep_rows.append(row_dict)

    epoch = getattr(rep_wb, 'epoch', None)
    rep_rows.sort(key=lambda r: _coerce_datetime(r.get('begin journey date'), epoch) or datetime.min, reverse=True)

    total = len(rep_rows)
    _report(progress_callback, 2, status='running', processed_rows=0, total_rows=total,
            message='Phase 2: Writing repJourney data into MAIN month sheet…')

    latest_date = None
    for row in rep_rows:
        candidate = _coerce_datetime(row.get('begin journey date'), epoch)
        if candidate:
            latest_date = candidate
            break

    month_sheet = _locate_month_sheet(main_wb, latest_date)
    _, month_header_map, month_header_row = _build_header_index(month_sheet, required_headers=['Destination'])

    destination_col = month_header_map.get('destination')
    if not destination_col:
        _report(progress_callback, 2, status='error', message="Phase 2 error: column 'Destination' not found in month sheet")
        raise PipelineError("Phase 2 error: column 'Destination' not found in month sheet", phase=2)

    shared_headers = [
        key for key in header_keys
        if key in month_header_map and month_header_map[key] < destination_col and key
    ]
    if not shared_headers:
        _report(progress_callback, 2, status='error', message='Phase 2 error: No overlapping headers between repJourney and month sheet before Destination column')
        raise PipelineError('Phase 2 error: No overlapping headers between repJourney and month sheet before Destination column', phase=2)

    data_start_row = month_header_row + 1
    original_last_row = month_sheet.max_row
    original_data_count = max(original_last_row - month_header_row, 0)
    formula_columns = list(range(destination_col, month_sheet.max_column + 1))
    formula_template = {}
    if original_data_count > 0:
        template_row_idx = data_start_row + original_data_count - 1
        for col in formula_columns:
            formula_template[col] = month_sheet.cell(row=template_row_idx, column=col).value

    for row_idx, row in enumerate(rep_rows):
        target_row = data_start_row + row_idx
        for header in shared_headers:
            col_idx = month_header_map[header]
            month_sheet.cell(row=target_row, column=col_idx, value=row.get(header))
        if (row_idx + 1) % 50 == 0 or (row_idx + 1) == total:
            _report(progress_callback, 2, processed_rows=row_idx + 1, total_rows=total,
                    message=f"Phase 2: repJourney merge – {row_idx + 1:,} / {total:,} rows")

    if original_data_count > len(rep_rows):
        rows_to_remove = original_data_count - len(rep_rows)
        month_sheet.delete_rows(data_start_row + len(rep_rows), rows_to_remove)

    if formula_columns and formula_template:
        for row_idx in range(original_data_count, len(rep_rows)):
            target_row = data_start_row + row_idx
            for col in formula_columns:
                month_sheet.cell(row=target_row, column=col, value=formula_template.get(col))

    _report(progress_callback, 2, status='done', processed_rows=total, total_rows=total,
            message='Phase 2 complete – repJourney data refreshed in latest month sheet.')

    return {
        'month_sheet': month_sheet,
        'month_header_map': month_header_map,
        'month_header_row': month_header_row,
        'formula_start_col': destination_col
    }


def _phase_three_update_main(main_wb, context, progress_callback):
    month_sheet: Worksheet = context['month_sheet']
    month_header_map = context['month_header_map']
    month_header_row = context.get('month_header_row', 1)

    month_device_col = _find_first_header(month_header_map, [
        'ivm/iscout device id', 'iscout device id', 'device nos', 'device_id', 'device id'
    ])
    disarm_col = month_header_map.get('disarm date')
    destination_col = month_header_map.get('destination')

    if not month_device_col or not disarm_col or not destination_col:
        _report(progress_callback, 3, status='error', message='Phase 3 error: Missing device, Disarm Date, or Destination columns in month sheet')
        raise PipelineError('Phase 3 error: Missing device, Disarm Date, or Destination columns in month sheet', phase=3)

    device_lookup: Dict[str, Tuple[object, object]] = {}
    for row in month_sheet.iter_rows(min_row=month_header_row + 1, max_row=month_sheet.max_row, values_only=False):
        device_value = row[month_device_col - 1].value
        if not device_value:
            continue
        device_key = _normalize_device_id(device_value)
        if not device_key or device_key in device_lookup:
            continue
        device_lookup[device_key] = (
            row[disarm_col - 1].value,
            row[destination_col - 1].value,
        )

    main_sheet = _locate_main_sheet(main_wb)
    _, main_header_map, main_header_row = _build_header_index(main_sheet, required_headers=['Device Nos', 'Device_ID', 'Device ID'])
    main_device_col = _find_first_header(main_header_map, ['device nos', 'device id', 'device_id'])
    last_disarmed_col = main_header_map.get('last disarmed date')
    last_area_col = main_header_map.get('last disarmed area')

    if not main_device_col or not last_disarmed_col or not last_area_col:
        _report(progress_callback, 3, status='error', message="Phase 3 error: MAIN sheet missing 'Device Nos', 'Last Disarmed Date', or 'Last Disarmed Area'")
        raise PipelineError("Phase 3 error: MAIN sheet missing 'Device Nos', 'Last Disarmed Date', or 'Last Disarmed Area'", phase=3)

    total_rows = max(main_sheet.max_row - main_header_row, 0)
    if total_rows == 0:
        _report(progress_callback, 3, status='done', message='Phase 3 complete – MAIN sheet contains no rows to update.')
        return

    _report(progress_callback, 3, status='running', processed_rows=0, total_rows=total_rows,
            message='Phase 3: Updating MAIN last disarmed fields…')

    for idx, row in enumerate(main_sheet.iter_rows(min_row=main_header_row + 1, max_row=main_sheet.max_row, values_only=False), start=1):
        device_value = row[main_device_col - 1].value
        if device_value is not None:
            device_key = _normalize_device_id(device_value)
            if device_key and device_key in device_lookup:
                last_date, last_area = device_lookup[device_key]
                row[last_disarmed_col - 1].value = last_date
                row[last_area_col - 1].value = last_area
        if idx % 50 == 0 or idx == total_rows:
            _report(progress_callback, 3, processed_rows=idx, total_rows=total_rows,
                    message=f"Phase 3: Updating MAIN last disarmed fields – {idx:,} / {total_rows:,} rows")

    _report(progress_callback, 3, status='done', processed_rows=total_rows, total_rows=total_rows,
            message='Phase 3 complete – MAIN sheet enriched with disarm details.')


def _build_header_index(
    sheet: Worksheet,
    required_headers: Optional[Iterable[str]] = None,
    max_scan_rows: int = 30,
) -> Tuple[List[str], Dict[str, int], int]:
    required = {_normalize_header(h) for h in required_headers} if required_headers else set()
    header_row_idx: Optional[int] = None
    headers: List[str] = []

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_scan_rows, max_col=sheet.max_column, values_only=False),
        start=1,
    ):
        values = [cell.value if cell.value is not None else '' for cell in row]
        normalized = [_normalize_header(value) for value in values]
        if required:
            if required.intersection(normalized):
                header_row_idx = row_idx
                headers = values
                break
        elif any(normalized):
            header_row_idx = row_idx
            headers = values
            break

    if header_row_idx is None:
        header_row_idx = 1
        row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
        headers = [cell.value if cell.value is not None else '' for cell in row]

    header_map: Dict[str, int] = {}
    for idx, title in enumerate(headers, start=1):
        key = _normalize_header(title)
        if key and key not in header_map:
            header_map[key] = idx

    return headers, header_map, header_row_idx


def _normalize_header(value) -> str:
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip().lower()
    return str(value).strip().lower()


def _normalize_device_id(value) -> str:
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        try:
            value = int(float(value))
            return str(value)
        except (ValueError, TypeError):
            pass
    text = str(value).strip()
    if not text:
        return ''
    if text.isdigit():
        text = str(int(text))
    stripped = text.lstrip('0')
    return stripped or '0'


def _coerce_datetime(value, epoch=None) -> Optional[datetime]:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            return excel_date(value, epoch)
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%Y %H:%M", "%d-%b-%Y"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def _locate_month_sheet(workbook, latest_date: Optional[datetime]) -> Worksheet:
    pattern = re.compile(r'^[A-Za-z]{3}\d{4}$')
    candidates = [name for name in workbook.sheetnames if pattern.match(name.strip())]
    if latest_date:
        target = latest_date.strftime('%b%Y')
        for name in candidates:
            if name.lower() == target.lower():
                return workbook[name]
    if candidates:
        return workbook[candidates[0]]
    raise PipelineError('Phase 2 error: No MMMYYYY month sheet found inside MAIN workbook.', phase=2)


def _locate_main_sheet(workbook) -> Worksheet:
    for name in workbook.sheetnames:
        if name.strip().lower() == 'main':
            return workbook[name]
    for sheet in workbook.worksheets:
        _, header_map, _ = _build_header_index(sheet)
        if 'device nos' in header_map or 'device id' in header_map:
            return sheet
    raise PipelineError("Phase 3 error: Unable to find MAIN sheet with 'Device Nos' column.", phase=3)


def _find_first_header(header_map: Dict[str, int], options: Iterable[str]) -> Optional[int]:
    for option in options:
        key = _normalize_header(option)
        if key in header_map:
            return header_map[key]
    return None


def _report(callback: Callable[..., None], phase: int, **payload):
    if callback:
        callback(phase, **payload)
